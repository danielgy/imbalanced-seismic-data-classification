from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import os
import pandas as pd

class op():
    def __init__(self):
        self.file_name = ""
        self.openst = 0
        self.closest = 0
        self.wb = ""
        self.ws = ""
        self.ew = ""
        self.strow = 1
    def openex(self):
        if os.path.exists(self.file_name):
            self.wb = load_workbook(self.file_name)
        else:
            self.wb = Workbook()

    def creatsht(self, shtname, shtindex):
        shtnum = len(self.wb.get_sheet_names())
        if shtindex <= shtnum:
            self.ws = self.wb.worksheets[shtnum - 1]
            self.ws.title = shtname
        else:
            # use ExcelWriter method to write file and save it
            self.ws = self.wb.create_sheet(shtname, shtindex)
    def closeex(self):
        self.wb.save(filename=self.file_name)
   
    def labelrow(self, df, rnum, cnum):
        rows = len(df)
        if isinstance(df,pd.DataFrame):
            if rows>0:
                for rx in range(rows):
                    self.ws.cell(row=rnum, column=rx + cnum).value = df.iloc[:,rx]
        elif isinstance(df,list):
            if rows>0:
                for rx in range(rows):
                    self.ws.cell(row=rnum, column=rx + cnum).value = df[rx]

    def labelcol(self, df, rnum, cnum):
        rows = len(df)
        if isinstance(df,pd.DataFrame):
            if rows>0:
                for rx in range(rows):
                    self.ws.cell(row=rx + rnum, column=cnum).value = df.iloc[rx,:]
        elif isinstance(df,list):
            if rows>0:
                for rx in range(rows):
                    self.ws.cell(row=rx + rnum, column=cnum).value = df[rx]
    def write_df(self,df):
        rows,cols=df.shape
        for ii in range(rows):
            self.ws.append(list(df.iloc[ii,:]))
            
    

